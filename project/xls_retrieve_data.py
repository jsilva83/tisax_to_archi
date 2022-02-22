# Import external packages.
import openpyxl
import os


class XlsRetrieveData:
    """This class reads the data from an Excel file."""

    def __init__(self, in_column: int, in_workbook: str, in_worksheet: str, in_first_row: int) -> None:
        """Opens the workbook and worksheet with AMS KPIs data."""
        # Open the xls file.
        self.first_row = in_first_row
        self.xls_work_book_obj = openpyxl.load_workbook(filename=in_workbook)
        self.xls_work_sheet_obj = self.xls_work_book_obj[in_worksheet]
        self.last_row = self.calculate_nr_of_rows(in_column)
        return

    def calculate_nr_of_rows(self, in_column) -> int:
        """Calculates the number of rows in the xls file."""
        this_row = self.first_row
        while True:
            cell_value = self.xls_work_sheet_obj.cell(row=this_row, column=in_column).value
            if cell_value is None:
                break
            this_row += 1
        return this_row - 1

    def get_column_data(self, in_column) -> list:
        """Returns a list with all fields of a column."""
        output_list = []
        for n in range(2, self.last_row + 1):
            if self.xls_work_sheet_obj.cell(row=n, column=in_column).value is None:
                cell_value = ''
            else:
                cell_value = self.xls_work_sheet_obj.cell(row=n, column=in_column).value
            output_list.append(cell_value)
        return output_list

    def get_must_row(self, in_row):
        out_row = []
        isa_cell = self.xls_work_sheet_obj.cell(row=in_row, column=4).value
        objective_cell = self.xls_work_sheet_obj.cell(row=in_row, column=10).value
        must_cell = self.xls_work_sheet_obj.cell(row=in_row, column=11).value
        if len(isa_cell.split('.')) == 3 and must_cell is not None:
            out_row.append(f'{isa_cell}. {objective_cell}')
            out_row.append(must_cell.split('\n'))
        return out_row

    def get_should_row(self, in_row):
        out_row = []
        isa_cell = self.xls_work_sheet_obj.cell(row=in_row, column=4).value
        objective_cell = self.xls_work_sheet_obj.cell(row=in_row, column=10).value
        should_cell = self.xls_work_sheet_obj.cell(row=in_row, column=12).value
        if len(isa_cell.split('.')) == 3 and should_cell is not None:
            out_row.append(f'{isa_cell}. {objective_cell}')
            out_row.append(should_cell.split('\n'))
        return out_row

    def get_high_row(self, in_row):
        out_row = []
        isa_cell = self.xls_work_sheet_obj.cell(row=in_row, column=4).value
        objective_cell = self.xls_work_sheet_obj.cell(row=in_row, column=10).value
        high_cell = self.xls_work_sheet_obj.cell(row=in_row, column=13).value
        if len(isa_cell.split('.')) == 3 and high_cell is not None:
            out_row.append(f'{isa_cell}. {objective_cell}')
            out_row.append(high_cell.split('\n'))
        return out_row

    def save_xls_files(self):
        return


if __name__ == '__main__':
    a_xls = XlsRetrieveData(4, 'C:/Users/jorge.silva/HUFGLOBAL/SGSI ISMS - ISMS Management - ISMS Management'
                                       '/I - TISAX Project/02 - Self assessment/05 - VDA 5_0_4 (05-01-2022)/VDA ISA '
                                       '5.0.4_EN_archimate.xlsx', 'Information Security', 5)
    for a_n in range(5, a_xls.last_row+1):
        a_row = a_xls.get_must_row(a_n)
        print(a_row)
